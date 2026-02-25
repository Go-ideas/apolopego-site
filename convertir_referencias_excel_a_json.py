from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


INPUT_FILE = Path("referencias_profesionales.xlsx")
OUTPUT_FILE = Path("referencias.json")
SHEET_NAME = "Referencias"


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_bool(value: Any) -> bool:
    text = _norm_text(value).lower()
    return text in {"si", "sí", "true", "1", "x", "yes"}


def _to_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {INPUT_FILE}")

    wb = load_workbook(INPUT_FILE, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{SHEET_NAME}' en {INPUT_FILE.name}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("La hoja está vacía.")

    headers = [_norm_text(h) for h in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}

    required = [
        "id",
        "nombre_completo",
        "cargo",
        "empresa",
        "relacion_profesional",
        "contexto_proyecto",
        "comentario",
        "autorizado_publicacion",
        "destacado",
    ]
    missing = [col for col in required if col not in header_index]
    if missing:
        raise ValueError(f"Faltan columnas requeridas en Excel: {', '.join(missing)}")

    references: list[dict[str, Any]] = []

    for excel_row in rows[1:]:
        if not excel_row or all(_norm_text(v) == "" for v in excel_row):
            continue

        comentario = _norm_text(excel_row[header_index["comentario"]])
        cargo = _norm_text(excel_row[header_index["cargo"]])
        empresa = _norm_text(excel_row[header_index["empresa"]])
        contexto = _norm_text(excel_row[header_index["contexto_proyecto"]])

        # Si no hay contenido mínimo, omitimos la fila.
        if not comentario and not (cargo or empresa or contexto):
            continue

        ref_id = _to_int(excel_row[header_index["id"]], len(references) + 1)
        nombre_completo = _norm_text(excel_row[header_index["nombre_completo"]])
        nombre_visible = (
            _norm_text(excel_row[header_index["nombre_visible"]])
            if "nombre_visible" in header_index
            else ""
        )
        anonimizar = (
            _to_bool(excel_row[header_index["anonimizar"]])
            if "anonimizar" in header_index
            else False
        )

        if anonimizar:
            nombre_completo = ""

        references.append(
            {
                "id": ref_id,
                "nombre_completo": nombre_completo,
                "nombre_visible": nombre_visible,
                "cargo": cargo,
                "empresa": "" if anonimizar else empresa,
                "relacion_profesional": _norm_text(
                    excel_row[header_index["relacion_profesional"]]
                ),
                "contexto_proyecto": contexto,
                "comentario": comentario,
                "autorizado_publicacion": _to_bool(
                    excel_row[header_index["autorizado_publicacion"]]
                ),
                "destacado": _to_bool(excel_row[header_index["destacado"]]),
            }
        )

    payload = {"references": references}
    OUTPUT_FILE.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"OK: {OUTPUT_FILE} generado con {len(references)} referencias.")


if __name__ == "__main__":
    main()
